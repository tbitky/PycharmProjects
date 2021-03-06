import os
import tkinter
import re
import pandas as pd
import xrayutilities as xu
import numpy as np
from tkinter import filedialog
from tkinter import messagebox
from scipy.signal import find_peaks, peak_widths
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
import subprocess
from time import sleep
from decimal import Decimal, ROUND_HALF_UP
from time import sleep
from tkinter import filedialog

import numpy as np
import openpyxl
import pandas as pd
import psutil
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from scipy import signal


def distinguish_sample_number_and_direction(file_name):
    """
    ファイル名からサンプル番号,測定方向を抽出。
    """
    line = re.compile(r"フィルタ\+\s傾斜G\d+_\d.xls")
    if re.match(line, file_name):
        nums = re.findall('\d+', file_name)
        sample_name = int(nums[0])
        direction = int(nums[1])
    else:
        sample_name, direction = file_name, 1

    return sample_name, direction


def fine_round(x, y=0):
    round_x = Decimal(str(x)).quantize(Decimal(str(y)), rounding=ROUND_HALF_UP)
    int_x = int(round_x)
    return int_x


def bowing_direction(data):
    horizon = (data[0] + data[-1]) / 2
    peak_cut = 3
    minimum_width=20
    max = 'max'
    min = 'min'
    left_index = [0, ]
    right_index = []
    direction = '-'
    greater_or_less = []
    for i in range(len(data) - 1):
        if (data[i] - horizon) * (data[i + 1] - horizon) < 0 and i - left_index[len(right_index)] >= minimum_width:
            right_index.append(i)
            if i<len(data)-2:
                left_index.append(i + 1)
    if len(right_index)<len(left_index):
        right_index.append(len(data) - 1)

    for i in range(len(left_index)):
        maxvalue = np.max(data[left_index[i]:right_index[i]])
        minvalue = np.min(data[left_index[i]:right_index[i]])
        if maxvalue - horizon >= peak_cut and maxvalue + minvalue >= 2 * horizon:
            greater_or_less.append(max)
        elif minvalue - horizon <= -1 * peak_cut and maxvalue + minvalue <= 2 * horizon:
            greater_or_less.append(min)

    if len(greater_or_less) == 1:
        if greater_or_less[0] == max:
            direction = '凸'
        elif greater_or_less[0] == min:
            direction = '凹'
    elif len(greater_or_less) == 2:
        direction = 'N'
    elif len(greater_or_less) == 3:
        if greater_or_less[0] == max:
            direction = 'M'
        elif greater_or_less[0] == min:
            direction = 'W'
    else:
        direction = False
    return direction


def main():
    """
    ファイルダイアログを開いてfilepathに格納
    """
    root = tkinter.Tk()
    root.withdraw()
    fTyp = [("excelファイル", "*.xls")]
    iDir = os.path.abspath(os.path.dirname(r"\\Sirius\carbon-nas\SR4000\001_bowing"))
    filepath = filedialog.askopenfilenames(filetypes=fTyp, initialdir=iDir)

    """
    ファイルをデータフレーム化
    """
    columns = ('shape_1', 'shape_2', 'bowing_1', 'bowing_2')
    Data = pd.DataFrame(columns=columns)
    for i in range(len(filepath)):
        dataws = pd.read_excel(filepath[i], sheet_name='測定結果')
        bowingdata = dataws.loc[:, 'Revised'].values
        shape = bowing_direction(bowingdata)
        height = dataws.iat[23, 5]
        file_name = os.path.basename(filepath[i])
        sample_name, direction = distinguish_sample_number_and_direction(file_name)
        Data.at[str(sample_name), 'sample_name'] = sample_name
        Data.at[str(sample_name), 'bowing_' + str(direction)] = height
        Data.at[str(sample_name), 'shape_' + str(direction)] = shape

    """
    Excelシートの初期化
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
    sheet.merge_cells(start_row=1, end_row=1, start_column=2, end_column=3)
    sheet.merge_cells(start_row=1, end_row=1, start_column=4, end_column=5)

    sheet['A1'].value = 'サンプル番号'
    sheet['B1'].value = '反り形状'
    sheet['D1'].value = '反り[um]'
    sheet['B2'].value = '1方向'
    sheet['C2'].value = '2方向'
    sheet['D2'].value = '1方向'
    sheet['E2'].value = '2方向'

    """
    シートに記入
    """
    minsample = int(np.min(Data.index.values))
    maxsample = int(np.max(Data.index.values))
    error = 0
    for i, j in enumerate(Data.index.values):
        if j.isdecimal():
            row = int(j) - minsample + 3
            sheet.cell(row=row, column=1).value = 'G' + str(j)
        else:
            error += 1
            row = maxsample - minsample + 3 + error
            sheet.cell(row=row, column=1).value = j
        for k, l in enumerate(columns):
            sheet.cell(row=row, column=k + 2).value = Data.at[j, l]

    """
    枠線とか整える
    """
    side = Side(style="thin", color="000000")
    border = Border(top=side, bottom=side, left=side, right=side)
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for col in sheet.iter_cols(min_col=1):
        for cell in col:
            cell.border = border
            cell.alignment = align
            if cell.column > 1:
                cell.number_format = '0.0'
    """
    ファイルを開く
    """
    user_directory = os.path.expanduser('~')
    excel_file = user_directory + '\AppData\Local\Temp\Bowingリスト.xlsx'
    wb.save(excel_file)
    subprocess.Popen([r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE", excel_file])

    """
    Excel閉じたらファイル削除
    """
    sleep(5)
    while True:
        flag = True
        ps = psutil.process_iter(attrs=['name'])
        for i, p in enumerate(ps):
            pi = p.info
            if pi['name'] == 'EXCEL.EXE':
                flag = False
        if flag:
            break
        sleep(1)
    os.remove(excel_file)


if __name__ == '__main__':
    main()
