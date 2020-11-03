import os
import tkinter
import re
import pandas as pd
import xrayutilities as xu
import numpy as np
from tkinter import filedialog
from tkinter import messagebox
from scipy.signal import find_peaks, peak_widths
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
import subprocess
from time import sleep
from decimal import Decimal, ROUND_HALF_UP
import psutil
import lib.egw.functions as egw


def distinguish_hkl_location_number(x):
    """
    ファイル名からサンプル番号、測定した結晶面、測定位置を抽出。
    """
    match_template = map(re.compile, [r"G\d+\sOmega\s\d{3}_upper_\d\.xrdml",
                                      r"G\d+\s1point_omega\s\d{3}\supper_\d\.xrdml"])
    for line in match_template:
        if re.match(line, x):
            nums = re.findall('\d+', x)
            sample_name = int(nums[0])
            plane = nums[len(nums) - 2]
            position = int(nums[len(nums) - 1])
            break
    else:
        sample_name, plane, position = None, None, None

    return sample_name, plane, position



def switching_fwhm_calculate(position, intensity, means='absolute'):
    peaks, properties = find_peaks(intensity, distance=len(intensity) // 2)
    if means == 'absolute':
        try:
            fwhm_in_degree = xu.math.misc.fwhm_exp(position, intensity)
        except:
            fwhm_in_degree = None
    elif means == 'relative':
        try:
            widths = peak_widths(intensity, peaks)
            left_edge = position[int(widths[2][0])] + (widths[2][0] - int(widths[2][0])) * (
                    position[int(widths[2][0]) + 1] - position[int(widths[2][0])])
            right_edge = position[int(widths[3][0])] + (widths[3][0] - int(widths[3][0])) * (
                    position[int(widths[3][0]) + 1] - position[int(widths[3][0])])
            fwhm_in_degree = np.abs(left_edge - right_edge)
        except TypeError:
            fwhm_in_degree = None
    if fwhm_in_degree:
        fwhm_in_arcsec = fwhm_in_degree * 3600
    else:
        fwhm_in_arcsec = fwhm_in_degree
    return fwhm_in_arcsec


def main():
    """
    ファイルダイアログを開いてfilepathに格納
    """
    root = tkinter.Tk()
    root.withdraw()
    fTyp = [("xrdファイル", "*.xrdml")]
    iDir = os.path.abspath(os.path.dirname(r"\\Sirius\carbon-nas\SR4000\003_XRD"))
    filepath = filedialog.askopenfilenames(filetypes=fTyp, initialdir=iDir)

    """
    ファイルをデータフレーム化
    """
    Data = pd.DataFrame(columns=('directory', 'file_name', 'sample_name', 'plane', 'position', 'FWHM'))
    for i in range(len(filepath)):
        file_name = os.path.basename(filepath[i])
        sample_name, plane, position = distinguish_hkl_location_number(file_name)
        Data.at[i, 'directory'] = os.path.dirname(filepath[i])
        Data.at[i, 'file_name'] = file_name
        Data.at[i, 'sample_name'] = sample_name
        Data.at[i, 'plane'] = plane
        Data.at[i, 'position'] = position
    """
    半値幅計算
    """
    for i in range(len(Data)):
        xrdfile = xu.io.panalytical_xml.XRDMLFile(Data.at[i, 'file_name'], path=Data.at[i, 'directory'])
        scanmot, intensity = (
            xu.io.panalytical_xml.getxrdml_scan(filetemplate=xrdfile.filename, motors=xrdfile.scan.scanmotname,
                                                path=os.path.dirname(xrdfile.full_filename)))
        count_time = xrdfile.scan.ddict['countTime']
        integer_intensity = np.array(intensity / count_time).astype(np.int)
        Data.at[i, 'FWHM'] = switching_fwhm_calculate(scanmot, integer_intensity, means='absolute')
    """
    Excelシートの初期化
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    for j in np.arange(1, 12):
        if j == 1:
            sheet.merge_cells(start_row=1, end_row=2, start_column=j, end_column=j)
        elif j == 2 or j == 6:
            sheet.merge_cells(start_row=1, end_row=1, start_column=j, end_column=j + 3)

    sheet["A1"].value = "サンプル番号"
    sheet["B1"].value = "002"
    sheet["F1"].value = "102"
    sheet["j1"].value = "100"
    for k in np.arange(1, 12):
        if k == 2 or k == 6:
            sheet.cell(row=2, column=k).value = "センター"
        elif k == 3 or k == 7:
            sheet.cell(row=2, column=k).value = "ミドル"
        elif k == 4 or k == 8:
            sheet.cell(row=2, column=k).value = "エッジ"
        elif k == 5 or k == 9:
            sheet.cell(row=2, column=k).value = "平均"
    """
    シートに記入
    """
    complete_data = ~(Data.isnull().any(axis=1))
    maxsample = np.max(Data.loc[complete_data, 'sample_name'])
    minsample = np.min(Data.loc[complete_data, 'sample_name'])
    error = 0
    for i in Data.index.values:
        if complete_data[i]:
            row = int(Data.at[i, 'sample_name']) - minsample + 3
            sheet.cell(row=row, column=1).value = int(Data.at[i, 'sample_name'])
            sheet.cell(row=row, column=5).value = '=AVERAGE(B:D)'
            sheet.cell(row=row, column=9).value = '=AVERAGE(E:I)'
            for j, k in enumerate(['002', '102', '100']):
                if Data.at[i, 'plane'] == k:
                    datacolumn = 2 + j * 4
            unique_positions = np.array(list((map(int, Data.loc[(Data['sample_name'] == Data.at[i, 'sample_name']) & (
                    Data['plane'] == Data.at[i, 'plane']), 'position'].unique()))))
            unique_positions = np.sort(unique_positions)
            for j, k in enumerate(unique_positions):
                if Data.at[i, 'position'] == k:
                    datacolumn += j
        '''
        不完全なデータは末尾に記入
        '''
        else:
            row = maxsample - minsample + 4 + error
            sheet.cell(row=row, column=1).value = Data.at[i, 'file_name']
            datacolumn = 2
            error += 1

        '''
        4000以上をX'pert epitaxyとの誤差注意とし、その場合色で警告
        '''
        if Data.at[i, 'FWHM'] is not None:
            sheet.cell(row=row, column=datacolumn).value = float(
                "{:.1f}".format(float(Data.at[i, 'FWHM'])))
            if float(Data.at[i, 'FWHM']) > 4000:  # 4000s以上を誤差注意とする
                fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFFF00", bgColor="FFFF00")
                sheet.cell(row=row, column=1).fill = fill

    """
    枠線とか整える
    """
    side = Side(style="thin", color="000000")
    border = Border(top=side, bottom=side, left=side, right=side)
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for col in sheet.iter_cols(min_col=1):
        for cell in col:
            cell.border = border
            cell.alignment = align
            if cell.column > 1:
                cell.number_format = '0.0'
    """
    ファイルを開く
    """
    user_directory = os.path.expanduser('~')
    excel_file = user_directory + '\AppData\Local\Temp\X線の半値幅リスト.xlsx'
    wb.save(excel_file)
    subprocess.Popen([r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE", excel_file])

    """
    Excel閉じたらファイル削除
    """
    sleep(5)
    while True:
        flag = True
        ps = psutil.process_iter(attrs=['name'])
        for i, p in enumerate(ps):
            pi = p.info
            if pi['name'] == 'EXCEL.EXE':
                flag = False
        if flag:
            break
        sleep(1)
    os.remove(excel_file)


if __name__ == '__main__':
    main()
